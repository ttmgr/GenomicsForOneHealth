#!/usr/bin/env python3
"""Add Quasimetagenomics + Metagenomics derived sheets to listeria_final_corrected.xlsx.

Idempotent: removes and recreates the two derived sheets on each run, leaving
the master ``Data`` sheet (307 rows, 56 cols) untouched.

Filters:
  Quasimetagenomics  -> Sample Type startswith 'quasimeta_' AND Swab Type == 'Sponge'
  Metagenomics       -> Sample Type == 'metagenomics'       AND Swab Type != '-'

The three T31/T32 rows with Swab Type == '-' are dropped from the derived
sheets (they have no swab association) but stay on the master Data sheet for
provenance.

Both derived sheets gain a trailing ``Timepoint`` column mapping Sample Type
to the canonical timeline:
  metagenomics      -> '0h (baseline)'
  quasimeta_4h      -> '4h'
  quasimeta_12h     -> '12h'
  quasimeta_24h_1   -> '24h (1)'
  quasimeta_24h_2   -> '24h (2)'
"""
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

XLSX = Path(__file__).resolve().parents[1] / "listeria_final_corrected.xlsx"
QUASI = "Quasimetagenomics"
META = "Metagenomics"

TIMEPOINT_MAP = {
    "metagenomics":    "0h (baseline)",
    "quasimeta_4h":    "4h",
    "quasimeta_12h":   "12h",
    "quasimeta_24h_1": "24h (1)",
    "quasimeta_24h_2": "24h (2)",
}
TIMEPOINT_SORT = {
    "0h (baseline)": 0, "4h": 1, "12h": 2, "24h (1)": 3, "24h (2)": 4,
}
SWAB_SORT = {"Cotton": 0, "Sponge": 1, "Zymo": 2}
TREATMENT_SORT = {"Lm2": 0, "Lm4": 1, "Lm6": 2, "Blank": 3, "Background": 4}


def _tp(values):
    return TIMEPOINT_MAP.get(values[0], "")


def _quasi_key(values):
    # col indices (0-based): SampleType=0, Condition=1, Treatment=2, LabID=3
    return (
        TIMEPOINT_SORT.get(_tp(values), 99),
        TREATMENT_SORT.get(values[2], 99),
        str(values[3]),
        values[1],
    )


def _meta_key(values):
    # Sort by swab (Cotton/Sponge/Zymo), treatment, lab id, condition
    return (
        SWAB_SORT.get(values[4], 99),
        TREATMENT_SORT.get(values[2], 99),
        str(values[3]),
        values[1],
    )


def main() -> None:
    wb = load_workbook(XLSX)
    data = wb["Data"]

    headers = [c.value for c in data[1]]
    n_cols = len(headers)

    # Collect data rows as (values tuple, source cells tuple). The source
    # cells stay valid for reading .number_format because we never mutate
    # the Data sheet.
    all_rows: list[tuple[tuple, tuple]] = []
    for row in data.iter_rows(min_row=2, max_row=data.max_row):
        values = tuple(c.value for c in row)
        if values[0] is None:
            continue
        all_rows.append((values, row))

    def quasi_filter(entry):
        v = entry[0]
        return isinstance(v[0], str) and v[0].startswith("quasimeta_") and v[4] == "Sponge"

    def meta_filter(entry):
        v = entry[0]
        return v[0] == "metagenomics" and v[4] != "-"

    quasi_rows = sorted((r for r in all_rows if quasi_filter(r)), key=lambda r: _quasi_key(r[0]))
    meta_rows = sorted((r for r in all_rows if meta_filter(r)), key=lambda r: _meta_key(r[0]))

    for name in (QUASI, META):
        if name in wb.sheetnames:
            del wb[name]

    def build_sheet(name: str, rows: list[tuple[tuple, tuple]]):
        ws = wb.create_sheet(name)

        # Header row: original headers + Timepoint column
        for col_idx, hval in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=hval)
            src = data.cell(row=1, column=col_idx)
            cell.font = copy(src.font)
            cell.fill = copy(src.fill)
            cell.alignment = copy(src.alignment)
            cell.border = copy(src.border)
        tp_col_idx = n_cols + 1
        tp_hdr = ws.cell(row=1, column=tp_col_idx, value="Timepoint")
        src_hdr = data.cell(row=1, column=1)
        tp_hdr.font = copy(src_hdr.font)
        tp_hdr.fill = copy(src_hdr.fill)
        tp_hdr.alignment = copy(src_hdr.alignment)
        tp_hdr.border = copy(src_hdr.border)

        # Data rows, preserving per-cell number formats
        for row_idx, (values, src_cells) in enumerate(rows, start=2):
            for col_idx, (val, src) in enumerate(zip(values, src_cells), start=1):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.number_format = src.number_format
            ws.cell(row=row_idx, column=tp_col_idx, value=_tp(values))

        # Inherit column widths from Data sheet
        for col_idx in range(1, n_cols + 1):
            letter = get_column_letter(col_idx)
            dim = data.column_dimensions.get(letter)
            if dim is not None and dim.width:
                ws.column_dimensions[letter].width = dim.width
        ws.column_dimensions[get_column_letter(tp_col_idx)].width = 14
        ws.freeze_panes = "A2"

    build_sheet(QUASI, quasi_rows)
    build_sheet(META, meta_rows)

    wb.save(XLSX)
    print(f"Saved {XLSX}")
    print(f"  Data            : {data.max_row - 1} rows (untouched)")
    print(f"  {QUASI}: {len(quasi_rows)} rows")
    print(f"  {META}    : {len(meta_rows)} rows")
    dropped = sum(1 for r in all_rows if r[0][4] == "-")
    print(f"  Dropped from derived sheets (Swab Type == '-'): {dropped}")


if __name__ == "__main__":
    main()
